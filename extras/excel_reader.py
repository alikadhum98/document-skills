"""
أداة قراءة البيانات من Excel وتحويلها إلى shared_data.py
التشغيل: python extras/excel_reader.py اسم_الملف.xlsx
"""

import sys, os, json, datetime
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
except ImportError:
    print("تثبيت openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    import openpyxl

# ===== قالب Excel المتوقع =====
EXCEL_TEMPLATE_STRUCTURE = """
الورقة 1: الانبعاثات
  A: السنة  |  B: الإجمالي (Gg CO2eq)  |  C: الطاقة  |  D: الصناعة  |  E: الزراعة  |  F: النفايات

الورقة 2: NDC
  A: المعامل  |  B: القيمة
  هدف_غير_مشروط_pct | 2.5
  هدف_مشروط_pct | 15.0
  ...

الورقة 3: التمويل
  A: المصدر  |  B: المبلغ (مليون $)
  GEF | 8.85
  GCF | 28.3
  ...
"""

def read_emissions_sheet(ws):
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            try:
                year = int(row[0])
                val  = float(row[1])
                data[year] = val
            except (TypeError, ValueError):
                continue
    return data

def read_ndc_sheet(ws):
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            key = str(row[0]).strip()
            try:
                data[key] = float(row[1])
            except (TypeError, ValueError):
                data[key] = row[1]
    return data

def read_finance_sheet(ws):
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            source = str(row[0]).strip()
            try:
                data[source] = float(row[1])
            except (TypeError, ValueError):
                continue
    return data

def update_shared_data(emissions=None, ndc=None, finance=None):
    """يحدّث core/shared_data.py بالبيانات الجديدة"""
    shared_path = os.path.join(ROOT, "core/shared_data.py")
    content     = open(shared_path, encoding='utf-8').read()
    import re

    if emissions:
        latest_year = max(emissions.keys())
        latest_val  = int(emissions[latest_year])

        # تحديث قيمة LATEST_YEAR
        content = re.sub(r'LATEST_YEAR\s*=\s*\d+',
                         f'LATEST_YEAR     = {latest_year}', content)

        # تحديث قيمة آخر سنة في EMISSIONS
        content = re.sub(
            rf'{latest_year}:\s*\d+',
            f'{latest_year}:       {latest_val}',
            content
        )
        print(f"  ✓  تحديث الانبعاثات: سنة {latest_year} = {latest_val:,} Gg")

    if finance:
        for source, amount in finance.items():
            key_map = {
                "GEF": "GEF_مليون",
                "GCF": "GCF_مليون",
                "ثنائي": "ثنائي_مليون",
            }
            if source in key_map:
                field = key_map[source]
                content = re.sub(
                    rf'"{field}":\s*[\d.]+',
                    f'"{field}":      {amount}',
                    content
                )
                print(f"  ✓  تحديث التمويل: {source} = ${amount}M")

    # تحديث تاريخ آخر تعديل
    content = re.sub(
        r'LAST_UPDATED = datetime\.date\.today\(\)\.isoformat\(\)',
        f'LAST_UPDATED = datetime.date.today().isoformat()',
        content
    )

    open(shared_path, 'w', encoding='utf-8').write(content)
    return True

def process_excel(filepath):
    if not os.path.exists(filepath):
        print(f"  ✗  الملف غير موجود: {filepath}")
        return False

    print(f"\n  قراءة: {os.path.basename(filepath)}")
    print("  " + "-"*40)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  ✗  خطأ في قراءة الملف: {e}")
        return False

    emissions = None
    ndc_data  = None
    finance   = None

    sheet_names = [s.lower() for s in wb.sheetnames]

    for name, ws in zip(wb.sheetnames, wb.worksheets):
        name_lower = name.lower()
        if any(x in name_lower for x in ["emission","انبعاث","ghg","inventory","جرد"]):
            emissions = read_emissions_sheet(ws)
            print(f"  ✓  ورقة الانبعاثات: {len(emissions)} سنة")
        elif any(x in name_lower for x in ["ndc","مساهم","هدف"]):
            ndc_data = read_ndc_sheet(ws)
            print(f"  ✓  ورقة NDC: {len(ndc_data)} معامل")
        elif any(x in name_lower for x in ["finance","تمويل","fund"]):
            finance = read_finance_sheet(ws)
            print(f"  ✓  ورقة التمويل: {len(finance)} مصدر")

    if not any([emissions, ndc_data, finance]):
        print("  ⚠  لم يُتعرَّف على أي ورقة. أسماء الأوراق المتوقعة:")
        print("     'الانبعاثات' أو 'Emissions'")
        print("     'NDC' أو 'المساهمة'")
        print("     'التمويل' أو 'Finance'")
        return False

    print("\n  تحديث shared_data.py...")
    update_shared_data(emissions, ndc_data, finance)

    # حفظ نسخة JSON من البيانات المستوردة
    out_json = os.path.join(ROOT, "extras",
               f"imported_{datetime.date.today()}.json")
    with open(out_json, 'w', encoding='utf-8') as f:
        json.dump({
            "المصدر":       os.path.basename(filepath),
            "التاريخ":      str(datetime.date.today()),
            "الانبعاثات":   {str(k): v for k,v in (emissions or {}).items()},
            "التمويل":      finance or {},
        }, f, ensure_ascii=False, indent=2)
    print(f"  ✓  نسخة JSON محفوظة: {os.path.basename(out_json)}")
    return True

def create_sample_template():
    """إنشاء ملف Excel نموذجي للتعبئة"""
    wb = openpyxl.Workbook()

    # ورقة الانبعاثات
    ws1 = wb.active
    ws1.title = "الانبعاثات"
    ws1.append(["السنة","الإجمالي (Gg CO2eq)","الطاقة","الصناعة","الزراعة","النفايات"])
    for year in range(2018, 2023):
        ws1.append([year, 190000, 165000, 9000, 12000, 4000])

    # ورقة NDC
    ws2 = wb.create_sheet("NDC")
    ws2.append(["المعامل","القيمة"])
    for row in [
        ["هدف_غير_مشروط_pct", 2.5],
        ["هدف_مشروط_pct",      15.0],
        ["احتياج_تمويل_مليار", 88.0],
    ]:
        ws2.append(row)

    # ورقة التمويل
    ws3 = wb.create_sheet("التمويل")
    ws3.append(["المصدر","المبلغ (مليون $)"])
    for row in [["GEF",8.85],["GCF",28.3],["ثنائي",650.0]]:
        ws3.append(row)

    tmpl_path = os.path.join(ROOT, "extras", "template_بيانات_المناخ.xlsx")
    wb.save(tmpl_path)
    print(f"  ✓  تم إنشاء النموذج: {tmpl_path}")
    return tmpl_path

if __name__ == "__main__":
    print("="*55)
    print("  أداة قراءة بيانات Excel")
    print("="*55)

    if len(sys.argv) < 2:
        print("\n  الاستخدام: python extras/excel_reader.py ملف.xlsx")
        print("\n  لإنشاء ملف نموذجي للتعبئة:")
        ans = input("  هل تريد إنشاء نموذج Excel؟ (y/n): ").strip().lower()
        if ans == 'y':
            create_sample_template()
    else:
        filepath = sys.argv[1]
        if not os.path.isabs(filepath):
            filepath = os.path.join(os.getcwd(), filepath)
        process_excel(filepath)
